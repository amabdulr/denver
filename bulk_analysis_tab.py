"""
Bulk Analysis Tab Component
============================
This module contains ALL functionality for the Bulk Analysis tab (Tab 3).

Purpose: 
    - Process multiple RCAs at once through ChapterFinder and ContentWriter workflows
    - Upload Excel files and select RCA column
    - Run RAG-enabled analysis on each row
    - Export results with new columns: ChapterFinder_Output, ContentWriter_Output, Processing_Status

Key Functions:
    - render_bulk_analysis_tab(): Main entry point called from streamlit_app.py
    - render_rca_section(): Handles file upload, column selection, processing controls
    - process_bulk_rca(): Main processing loop with progress tracking
    - show_processing_results(): Display results and enable downloads
    - create_output_excel(): Merge results with original Excel data

Architecture:
    streamlit_app.py (Tab 3) → render_bulk_analysis_tab() → All logic contained here
"""

import streamlit as st
import pandas as pd
import io
import time
from datetime import datetime
from app_functions import run_agent_with_prompt_file


def render_bulk_analysis_tab(product_name: str):
    """Render the Bulk Analysis tab content"""
    st.markdown("---")
    st.markdown("### 📊 Bulk RCA Processing")
    st.warning("🧪 **Testing Mode Active:** Processing limited to first 10 rows only")
    st.info("💡 Process multiple RCAs at once through ChapterFinder and ContentWriter workflows")
    
    # Create two columns: RCA (active) and Bug (placeholder)
    col_rca, col_bug = st.columns(2)
    
    with col_rca:
        st.markdown("#### 📝 RCA Section")
        render_rca_section(product_name)
    
    with col_bug:
        st.markdown("#### 🐛 Bug Section")
        st.info("🚧 Coming soon - Bulk bug analysis functionality")
        st.markdown("*This section will allow bulk processing of bug IDs*")


def render_rca_section(product_name: str):
    """Render the RCA section with file upload and processing"""
    
    # Important note about Streamlit behavior
    if 'bulk_df' not in st.session_state:
        st.info("📌 **Note:** After uploading a file, Streamlit will refresh. Simply click back on the 'Bulk Analysis' tab to continue.")
    
    # File uploader with unique key
    uploaded_file = st.file_uploader(
        "Upload Excel file containing RCAs",
        type=['xlsx', 'xls'],
        help="Upload an Excel file with a column containing RCA text",
        key="bulk_excel_uploader"
    )
    
    if uploaded_file is not None:
        try:
            # Read the Excel file only if it's a new file
            if 'bulk_df' not in st.session_state or st.session_state.get('bulk_file_name') != uploaded_file.name:
                df = pd.read_excel(uploaded_file)
                st.session_state.bulk_df = df
                st.session_state.bulk_file_name = uploaded_file.name
                st.session_state.bulk_processed_rows = []
                st.session_state.bulk_processing_complete = False
            else:
                # Use cached dataframe
                df = st.session_state.bulk_df
            
            st.success(f"✅ File loaded: {uploaded_file.name} ({len(df)} rows, {len(df.columns)} columns)")
            
            # Show preview
            with st.expander("👀 Preview first 5 rows", expanded=True):
                st.dataframe(df.head(), use_container_width=True)
            
            # Column selection
            st.markdown("##### Select RCA Column")
            columns = df.columns.tolist()
            
            # Smart default: Try to find column with 'rca' in name
            default_idx = 0
            for idx, col in enumerate(columns):
                if 'rca' in str(col).lower():
                    default_idx = idx
                    break
            
            selected_column = st.selectbox(
                "Which column contains the RCA text?",
                options=columns,
                index=default_idx,
                help="Select the column that contains RCA descriptions"
            )
            
            # Validation
            if selected_column:
                non_empty = df[selected_column].notna().sum()
                st.info(f"📊 Column '{selected_column}' has {non_empty} non-empty cells out of {len(df)} rows")
                
                # Show sample from selected column
                with st.expander("📄 Sample RCA text from selected column"):
                    sample_texts = df[selected_column].dropna().head(3)
                    for idx, text in enumerate(sample_texts, 1):
                        st.markdown(f"**Sample {idx}:**")
                        st.text(str(text)[:300] + ("..." if len(str(text)) > 300 else ""))
                        st.markdown("---")
                
                # Processing controls
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 1, 1])
                
                with col1:
                    # Determine button text based on state
                    has_partial_results = len(st.session_state.get('bulk_processed_rows', [])) > 0
                    is_complete = st.session_state.get('bulk_processing_complete', False)
                    
                    if has_partial_results and not is_complete:
                        button_label = "▶️ Resume Processing"
                    else:
                        button_label = "🚀 Start Processing"
                    
                    process_button = st.button(
                        button_label,
                        type="primary",
                        use_container_width=True,
                        disabled=st.session_state.get('bulk_processing', False),
                        key="bulk_process_button"
                    )
                
                with col2:
                    if st.session_state.get('bulk_processing', False):
                        stop_button = st.button(
                            "⏸️ Stop",
                            type="secondary",
                            use_container_width=True,
                            key="bulk_stop_button"
                        )
                        if stop_button:
                            st.session_state.bulk_stop_requested = True
                
                with col3:
                    clear_button = st.button(
                        "🗑️ Clear",
                        use_container_width=True,
                        key="bulk_clear_button"
                    )
                    if clear_button:
                        clear_bulk_session_state()
                        st.rerun()
                
                # Time estimate (for first 10 rows in testing mode)
                if not st.session_state.get('bulk_processing', False):
                    rows_to_process = min(10, len(df))
                    estimated_time = calculate_processing_time(rows_to_process)
                    st.info(f"⏱️ Estimated processing time for first {rows_to_process} rows (testing mode): {estimated_time}")
                
                # Process the data
                if process_button:
                    if not product_name or product_name == "Select a product":
                        st.error("⚠️ Please select a product before processing")
                    else:
                        # Set processing flag IMMEDIATELY to prevent double-clicks
                        st.session_state.bulk_processing = True
                        st.session_state.bulk_stop_requested = False
                        process_bulk_rca(df, selected_column, product_name)
                
                # Show results if processing is complete or in progress
                if st.session_state.get('bulk_processed_rows'):
                    show_processing_results()
                
        except Exception as e:
            st.error(f"❌ Error reading Excel file: {str(e)}")
            with st.expander("🐛 Error Details"):
                st.exception(e)
    else:
        st.info("📤 Upload an Excel file to begin")


def calculate_processing_time(num_rows: int) -> str:
    """Calculate estimated processing time based on rate limiting"""
    # Each row requires 2 API calls (ChapterFinder + ContentWriter)
    # Rate limit: ~4.5 seconds per call
    # So: 2 calls × 4.5s = 9 seconds per row
    seconds_per_row = 9
    total_seconds = num_rows * seconds_per_row
    
    if total_seconds < 60:
        return f"{total_seconds} seconds"
    elif total_seconds < 3600:
        minutes = total_seconds / 60
        return f"{minutes:.1f} minutes"
    else:
        hours = total_seconds / 3600
        return f"{hours:.1f} hours"


def process_bulk_rca(df: pd.DataFrame, rca_column: str, product_name: str):
    """Process all RCAs in the DataFrame (limited to first 10 rows for testing)"""
    # Note: bulk_processing flag is set by caller before this function
    
    # Initialize results if starting fresh
    if not st.session_state.get('bulk_processed_rows'):
        st.session_state.bulk_processed_rows = []
        st.session_state.bulk_start_row = 0
    
    # Create progress indicators
    progress_bar = st.progress(0)
    status_text = st.empty()
    results_container = st.container()
    
    # **TESTING MODE: Limit to first 10 rows**
    total_rows = min(10, len(df))
    st.info(f"🧪 **Testing Mode:** Processing limited to first {total_rows} rows (out of {len(df)} total)")
    
    # Determine starting point (for resume functionality)
    start_idx = len(st.session_state.bulk_processed_rows)
    
    # Process each row
    for idx in range(start_idx, total_rows):
        # Check for stop request
        if st.session_state.get('bulk_stop_requested', False):
            st.warning("⏸️ Processing stopped by user")
            break
        
        row = df.iloc[idx]
        status_text.text(f"🔄 Processing row {idx + 1} of {total_rows}...")
        
        try:
            # Get RCA text
            rca_text = str(row[rca_column]) if pd.notna(row[rca_column]) else ""
            
            if not rca_text or rca_text.strip() == "" or rca_text == "nan":
                # Empty cell
                result = {
                    'row_number': idx + 1,
                    'rca_text_preview': "(empty)",
                    'chapter_finder': "",
                    'content_writer': "",
                    'status': 'Skipped - Empty cell'
                }
            else:
                # Process ChapterFinder
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Running ChapterFinder...")
                chapter_output = run_agent_with_prompt_file(
                    "ChapterFinder.md",
                    rca_text,
                    product_name
                )
                
                # Rate limiting: Wait 10 seconds between API calls (conservative to avoid rate limits)
                time.sleep(10)
                
                # Process ContentWriter
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Running ContentWriter...")
                content_output = run_agent_with_prompt_file(
                    "ContentWriter.md",
                    rca_text,
                    product_name
                )
                
                # Rate limiting: Wait 10 seconds after last call before next row
                time.sleep(10)
                
                result = {
                    'row_number': idx + 1,
                    'rca_text_preview': rca_text[:100] + "..." if len(rca_text) > 100 else rca_text,
                    'chapter_finder': chapter_output,
                    'content_writer': content_output,
                    'status': 'Success ✅'
                }
            
        except Exception as e:
            # Check error type
            error_str = str(e).lower()
            
            # Rate limit error (429)
            if '429' in error_str or 'rate limit' in error_str or 'spike arrest' in error_str:
                error_msg = f'Rate Limit Error: Too many API calls. System will add delays automatically.'
            # Network error
            elif any(term in error_str for term in ['connection', 'network', 'timeout', 'unreachable', 'failed to establish']):
                error_msg = f'Network Error: {str(e)[:100]} - Can resume when connection is restored'
            # Other errors
            else:
                error_msg = f'Error: {str(e)[:100]}'
            
            result = {
                'row_number': idx + 1,
                'rca_text_preview': str(row[rca_column])[:100] if pd.notna(row[rca_column]) else "(empty)",
                'chapter_finder': "",
                'content_writer': "",
                'status': error_msg
            }
            
            # If rate limit or network error, show alert
            if 'Rate Limit' in error_msg:
                st.warning(f"⚠️ Rate limit hit at row {idx + 1}. The system now includes 4.5s delays to prevent this. Clear results and restart for best results.")
            elif 'Network Error' in error_msg:
                st.error(f"⚠️ Network connection lost at row {idx + 1}. Your progress has been saved. Restore connection and click Resume.")
                st.session_state.bulk_stop_requested = True
        
        # Store result
        st.session_state.bulk_processed_rows.append(result)
        
        # Update progress
        progress = (idx + 1) / total_rows
        progress_bar.progress(progress)
    
    # Mark as complete
    st.session_state.bulk_processing = False
    if not st.session_state.get('bulk_stop_requested', False):
        st.session_state.bulk_processing_complete = True
        status_text.text(f"✅ Processing complete! Processed {len(st.session_state.bulk_processed_rows)} rows")
    else:
        status_text.text(f"⏸️ Processing paused at row {len(st.session_state.bulk_processed_rows)}")


def show_processing_results():
    """Display the processed results in a table and provide download"""
    st.markdown("---")
    st.markdown("#### 📊 Processing Results")
    
    results = st.session_state.bulk_processed_rows
    
    # Summary stats
    total = len(results)
    successful = sum(1 for r in results if r['status'] == 'Success ✅')
    skipped = sum(1 for r in results if 'Skipped' in r['status'])
    errors = sum(1 for r in results if 'Error' in r['status'])
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Processed", total)
    col2.metric("Successful", successful)
    col3.metric("Skipped", skipped)
    col4.metric("Errors", errors)
    
    # Results preview
    with st.expander("👀 Preview Results", expanded=True):
        preview_df = pd.DataFrame(results)
        st.dataframe(preview_df, use_container_width=True, height=400)
    
    # Prepare download
    st.markdown("#### 💾 Download Results")
    
    # Create output DataFrame by merging with original
    output_df = create_output_excel(results)
    
    # Convert to Excel with proper settings
    excel_buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl', mode='w') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Processed RCAs')
            
            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Processed RCAs']
            
            # Set column widths for better readability
            for idx, col in enumerate(output_df.columns, 1):
                # Make output columns wider
                if 'Output' in col or 'Status' in col:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 50
                else:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 20
            
            # Enable text wrapping for all cells
            from openpyxl.styles import Alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error creating Excel file: {str(e)}")
        st.info("💡 Trying simplified Excel export without formatting...")
        
        # Fallback: Create simple Excel without formatting
        excel_buffer = io.BytesIO()
        output_df.to_excel(excel_buffer, index=False, sheet_name='Processed RCAs', engine='openpyxl')
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
    
    # Download button
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bulk_rca_processed_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Download Processed Excel",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="bulk_download_button"
    )
    
    if st.session_state.get('bulk_processing_complete'):
        st.success("✅ All rows processed successfully! You can download the results above.")
    else:
        processed_count = len(st.session_state.get('bulk_processed_rows', []))
        st.info(f"ℹ️ Processing incomplete ({processed_count} rows processed). You can download partial results above and click '▶️ Resume Processing' to continue.")


def create_output_excel(results: list) -> pd.DataFrame:
    """Create the output Excel with original data + new columns"""
    original_df = st.session_state.bulk_df.copy()
    
    # Create results dataframe
    results_df = pd.DataFrame(results)
    
    # Sanitize text fields to prevent Excel corruption
    def sanitize_text(text):
        """Clean text for Excel compatibility"""
        if pd.isna(text) or text == "":
            return ""
        text = str(text)
        # Remove null characters and other problematic characters
        text = text.replace('\x00', '').replace('\r', '\n')
        # Limit length to prevent Excel cell limit issues (32,767 characters)
        if len(text) > 32000:
            text = text[:32000] + "\n\n[Content truncated due to length]"
        return text
    
    # Apply sanitization to output columns
    if 'chapter_finder' in results_df.columns:
        results_df['chapter_finder'] = results_df['chapter_finder'].apply(sanitize_text)
    if 'content_writer' in results_df.columns:
        results_df['content_writer'] = results_df['content_writer'].apply(sanitize_text)
    
    # Add new columns to original dataframe
    original_df['Row_Number'] = range(1, len(original_df) + 1)
    
    # Merge results (left join to keep all original rows)
    merged_df = original_df.merge(
        results_df[['row_number', 'chapter_finder', 'content_writer', 'status']],
        left_on='Row_Number',
        right_on='row_number',
        how='left'
    )
    
    # Rename columns
    merged_df = merged_df.rename(columns={
        'chapter_finder': 'ChapterFinder_Output',
        'content_writer': 'ContentWriter_Output',
        'status': 'Processing_Status'
    })
    
    # Drop temporary row number column
    merged_df = merged_df.drop(columns=['row_number', 'Row_Number'])
    
    # Fill NaN in new columns (for unprocessed rows)
    merged_df['ChapterFinder_Output'] = merged_df['ChapterFinder_Output'].fillna('Not processed')
    merged_df['ContentWriter_Output'] = merged_df['ContentWriter_Output'].fillna('Not processed')
    merged_df['Processing_Status'] = merged_df['Processing_Status'].fillna('Not processed')
    
    return merged_df


def clear_bulk_session_state():
    """Clear all bulk analysis session state"""
    keys_to_clear = [
        'bulk_df',
        'bulk_file_name',
        'bulk_processed_rows',
        'bulk_processing',
        'bulk_processing_complete',
        'bulk_stop_requested',
        'bulk_start_row'
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
