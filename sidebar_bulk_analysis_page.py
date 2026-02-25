"""
Bulk Analysis Sidebar Page Component
Handles the Bulk Analysis workflow for sidebar navigation app
"""

import streamlit as st
import pandas as pd
import io
import re
import time
from datetime import datetime
from app_functions import run_agent_with_prompt_file, run_agent, match_terms_to_guides, extract_doc_clues_data, load_document_inventory


def render_bulk_analysis_page():
    """Render the Bulk Analysis page with full functionality for sidebar navigation"""
    st.header("📊 Bulk Analysis")
    st.markdown("Process multiple RCAs at once through ChapterFinder and ContentWriter workflows")
    
    st.markdown("---")
    st.warning("🧪 **Testing Mode Active:** Processing limited to first 2 rows only")
    st.info("💡 Process multiple RCAs at once through ChapterFinder and ContentWriter workflows")
    
    # Step-by-step instructions
    with st.expander("📖 How to Use This Page", expanded=False):
        st.markdown("""
        #### RCA Section (Left Column)
        1. **Select Product**: Choose your product from the dropdown below
        2. **Upload Excel File**: Upload an Excel file containing RCA text
        3. **Select RCA Column**: Choose which column contains your RCA descriptions. A column is selected automatically.
        4. **Review Preview**: Check the preview to ensure correct column selection
        5. **Start Processing**: Click "🚀 Start Processing" to begin
        6. **Download Results**: Once complete, download the processed Excel file with Top-3 Recommendations and AI Analysis.
        
        #### Bug Section (Right Column)
        1. **Select Product**: Use the same product selection from below
        2. **Choose Note Extraction**: Check the box if you want to extract all notes (default: Behavior-changed + Release-note only)
        3. **Upload Excel File**: Upload an Excel file containing bug numbers (format: CSCxxxxxxx)
        4. **Select Bug Column**: The system will auto-detect columns with CSC-format bugs
        5. **Review Preview**: Verify the correct bug numbers are showing
        6. **Start Processing**: Click "🚀 Start Processing" to fetch bugs from CDETS and analyze them
        7. **Download Results**: Once complete, download the processed Excel with Top-3 Recommendation and BugAnalyze outputs
        
        **Note**: You can pause processing anytime and resume later. Your progress is saved!
        """)
    
    st.markdown("---")
    
    # Product name selection for bulk analysis
    from sidebar_app import get_saved_product, save_product_preference
    
    product_options = ["Cisco SD-WAN", "Cisco 9800", "ASR 9000", "Cisco 8000", "cisco_generic"]
    saved_product = get_saved_product()
    
    # Find index of saved product, default to 1 if not found
    try:
        default_index = product_options.index(saved_product) if saved_product in product_options else 1
    except ValueError:
        default_index = 1
    
    bulk_product_name = st.selectbox(
        "Select Product (will be saved for future use)",
        options=product_options,
        index=default_index,
        key='sidebar_bulk_product_selector'
    )
    save_product_preference(bulk_product_name)
    
    # ===== GUIDE SELECTION SECTION =====
    from sidebar_app import get_available_guides
    
    st.markdown("### 📚 Select Guides")
    st.caption("Limit the search scope to specific guides (optional)")
    
    # Get available guides for the selected product
    available_guides = get_available_guides(bulk_product_name)
    
    if available_guides:
        # Define default guides for Cisco SD-WAN (curated subset)
        sdwan_default_guides = [
            "systems-interfaces-book-xe-sdwan.pdf",
            "security-book-xe.pdf",
            "sdwan-xe-gs-book.pdf",
            "policies-book-xe.pdf",
            "appqoe-book-xe.pdf",
            "cloud-onramp-book-xe.pdf",
            "monitor-maintain-book.pdf",
            "compatibility-and-server-recommendations.pdf"
        ]
        
        # Initialize session state for selected guides if not exists
        # Default: all guides selected (except for SD-WAN which uses curated subset)
        if 'bulk_selected_guides' not in st.session_state:
            if bulk_product_name == "Cisco SD-WAN":
                # For SD-WAN, only select the curated guides that exist in available_guides
                st.session_state.bulk_selected_guides = [g for g in sdwan_default_guides if g in available_guides]
            else:
                st.session_state.bulk_selected_guides = available_guides.copy()
        
        # Also reset guides when product changes
        if 'bulk_last_product' not in st.session_state or st.session_state.bulk_last_product != bulk_product_name:
            if bulk_product_name == "Cisco SD-WAN":
                # For SD-WAN, only select the curated guides that exist in available_guides
                st.session_state.bulk_selected_guides = [g for g in sdwan_default_guides if g in available_guides]
            else:
                st.session_state.bulk_selected_guides = available_guides.copy()
            st.session_state.bulk_last_product = bulk_product_name
        
        # Add "Select All" / "Deselect All" buttons
        col_guide1, col_guide2 = st.columns(2)
        with col_guide1:
            if st.button("✅ Select All", use_container_width=True, key="bulk_select_all_guides"):
                st.session_state.bulk_selected_guides = available_guides.copy()
                # Clear all checkbox widget states to force refresh
                for guide in available_guides:
                    key = f"bulk_guide_{guide}"
                    if key in st.session_state:
                        st.session_state[key] = True
                st.rerun()
        with col_guide2:
            if st.button("❌ Deselect All", use_container_width=True, key="bulk_deselect_all_guides"):
                st.session_state.bulk_selected_guides = []
                # Clear all checkbox widget states to force refresh
                for guide in available_guides:
                    key = f"bulk_guide_{guide}"
                    if key in st.session_state:
                        st.session_state[key] = False
                st.rerun()
        
        # Display guides as checkboxes in an expander
        with st.expander(f"📖 Available Guides ({len(available_guides)})", expanded=True):
            st.caption(f"Found {len(available_guides)} guide(s) for {bulk_product_name}")
            
            # Create checkboxes for each guide
            for guide in available_guides:
                # Initialize checkbox state if not exists
                checkbox_key = f"bulk_guide_{guide}"
                if checkbox_key not in st.session_state:
                    st.session_state[checkbox_key] = guide in st.session_state.bulk_selected_guides
                
                st.checkbox(guide, key=checkbox_key)
            
            # Collect selected guides from checkboxes
            selected_guides = [guide for guide in available_guides if st.session_state.get(f"bulk_guide_{guide}", False)]
            st.session_state.bulk_selected_guides = selected_guides
            
            # Show selection summary
            if selected_guides:
                st.success(f"✅ {len(selected_guides)} guide(s) selected")
            else:
                st.info("ℹ️ No guides selected - will search all guides")
    else:
        st.warning(f"⚠️ No guides found for {bulk_product_name}")
    
    # Call bulk analysis content function
    render_bulk_analysis_content(bulk_product_name)


def render_bulk_analysis_content(product_name: str):
    """Render the Bulk Analysis tab content"""
    st.markdown("---")
    
    # Create two columns: RCA (active) and Bug (placeholder)
    col_rca, col_bug = st.columns(2)
    
    with col_rca:
        st.markdown("#### 📝 RCA Section")
        render_rca_section(product_name)
    
    with col_bug:
        st.markdown("#### 🐛 Bug Section")
        render_bug_section(product_name)
    
    # Test Section
    st.markdown("---")
    st.subheader("🧪 Capture your test results!")
    
    with st.expander("📝 Test Results", expanded=False):
        st.markdown("Capture test results for bulk analysis")
        
        # Feature being tested
        bulk_test_feature = st.text_input(
            "Feature to be tested",
            placeholder="e.g., bug analysis, bug summarize, bulk RCA, bulk bugs, First draft, resolve bug",
            help="Enter the specific feature or functionality being tested",
            key="bulk_test_feature"
        )
        
        # Name of tester
        bulk_tester_name = st.text_input(
            "Name of tester",
            placeholder="Enter your name",
            help="Name of the person performing the test",
            key="bulk_tester_name"
        )
        
        # Sliders for accuracy ratings
        bulk_location_accuracy = st.slider(
            "Location Accuracy",
            min_value=1,
            max_value=10,
            value=10,
            help="Rate the accuracy of the location/chapter recommendations (1=Poor, 10=Excellent)",
            key="bulk_test_location_accuracy"
        )
        
        bulk_content_accuracy = st.slider(
            "Content Accuracy",
            min_value=1,
            max_value=10,
            value=10,
            help="Rate the accuracy of the content generated (1=Poor, 10=Excellent)",
            key="bulk_test_content_accuracy"
        )
        
        # Comments text area
        bulk_test_comments = st.text_area(
            "Comments",
            placeholder="Enter any additional comments or observations...",
            height=100,
            key="bulk_test_comments"
        )
        
        # Wishlist text area
        bulk_test_wishlist = st.text_area(
            "Wishlist",
            placeholder="Enter feature requests, improvements, or wishlist items...",
            height=100,
            key="bulk_test_wishlist"
        )
        
        # Usefulness Rating
        st.markdown("---")
        bulk_usefulness = st.radio(
            "How useful is this feature?",
            options=[
                "⛔ I'd rather do this without AI",
                "🤔 Neutral - No strong preference",
                "👍 Yes, this is useful",
                "⭐ I'd prefer CIRCUIT over manual work"
            ],
            index=2,
            help="Rate how useful you find this AI-assisted feature",
            key="bulk_usefulness_rating"
        )
        
        # Add to Excel button
        bulk_add_to_excel_button = st.button(
            "📊 Add to Test Excel",
            type="primary",
            use_container_width=True,
            key="bulk_add_to_test_excel"
        )
        
        if bulk_add_to_excel_button:
            # Import the function from sidebar_app
            from sidebar_app import save_test_results_to_excel
            import os
            
            # Save to Excel with N/A for bug number and output content
            try:
                save_test_results_to_excel(
                    page_name="Bulk Analysis",
                    feature=bulk_test_feature,
                    tester_name=bulk_tester_name,
                    bug_number="N/A",
                    output_content="N/A",
                    location_accuracy=bulk_location_accuracy,
                    content_accuracy=bulk_content_accuracy,
                    comments=bulk_test_comments,
                    wishlist=bulk_test_wishlist,
                    usefulness=bulk_usefulness
                )
                st.success("✅ Test results saved to testresults.xlsx!")
                
                # Provide download link
                try:
                    with open("testresults.xlsx", "rb") as file:
                        st.download_button(
                            label="📥 Download testresults.xlsx",
                            data=file,
                            file_name="testresults.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="bulk_download_test_results"
                        )
                except:
                    pass
            except Exception as e:
                st.error(f"❌ Error saving to Excel: {str(e)}")
                with st.expander("🐛 Error Details"):
                    st.exception(e)


def render_rca_section(product_name: str):
    """Render the RCA section with file upload and processing"""
    
    # File uploader with unique key
    uploaded_file = st.file_uploader(
        "Upload Excel file containing RCAs",
        type=['xlsx', 'xls'],
        help="Upload an Excel file with a column containing RCA text",
        key="sidebar_bulk_excel_uploader"
    )
    
    if uploaded_file is not None:
        try:
            # Read the Excel file only if it's a new file
            if 'bulk_df' not in st.session_state or st.session_state.get('bulk_file_name') != uploaded_file.name:
                df = pd.read_excel(uploaded_file)
                st.session_state.bulk_df = df
                st.session_state.bulk_file_name = uploaded_file.name
                st.session_state.bulk_processed_rows = []
                st.session_state.bulk_processing_complete = False
            else:
                # Use cached dataframe
                df = st.session_state.bulk_df
            
            st.success(f"✅ File loaded: {uploaded_file.name} ({len(df)} rows, {len(df.columns)} columns)")
            
            # Show preview
            with st.expander("👀 Preview first 5 rows", expanded=True):
                st.dataframe(df.head(), use_container_width=True)
            
            # Column selection
            st.markdown("##### Select RCA Column")
            columns = df.columns.tolist()
            
            # Smart default: Try to find column with 'rca' in name
            default_idx = 0
            for idx, col in enumerate(columns):
                if 'rca' in str(col).lower():
                    default_idx = idx
                    break
            
            selected_column = st.selectbox(
                "Which column contains the RCA text?",
                options=columns,
                index=default_idx,
                help="Select the column that contains RCA descriptions",
                key="sidebar_bulk_column_selector"
            )
            
            # Validation
            if selected_column:
                non_empty = df[selected_column].notna().sum()
                st.info(f"📊 Column '{selected_column}' has {non_empty} non-empty cells out of {len(df)} rows")
                
                # Show sample from selected column
                with st.expander("📄 Sample RCA text from selected column"):
                    sample_texts = df[selected_column].dropna().head(3)
                    for idx, text in enumerate(sample_texts, 1):
                        st.markdown(f"**Sample {idx}:**")
                        st.text(str(text)[:300] + ("..." if len(str(text)) > 300 else ""))
                        st.markdown("---")
                
                # Processing controls
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 1, 1])
                
                with col1:
                    # Determine button text based on state
                    has_partial_results = len(st.session_state.get('bulk_processed_rows', [])) > 0
                    is_complete = st.session_state.get('bulk_processing_complete', False)
                    
                    if has_partial_results and not is_complete:
                        button_label = "▶️ Resume Processing"
                    else:
                        button_label = "🚀 Start Processing"
                    
                    process_button = st.button(
                        button_label,
                        type="primary",
                        use_container_width=True,
                        disabled=st.session_state.get('bulk_processing', False),
                        key="sidebar_bulk_process_button"
                    )
                
                with col2:
                    if st.session_state.get('bulk_processing', False):
                        stop_button = st.button(
                            "⏸️ Stop",
                            type="secondary",
                            use_container_width=True,
                            key="sidebar_bulk_stop_button"
                        )
                        if stop_button:
                            st.session_state.bulk_stop_requested = True
                
                with col3:
                    clear_button = st.button(
                        "🗑️ Clear",
                        use_container_width=True,
                        key="sidebar_bulk_clear_button"
                    )
                    if clear_button:
                        clear_bulk_session_state()
                        st.rerun()
                
                # Time estimate (for first 40 rows in testing mode)
                if not st.session_state.get('bulk_processing', False):
                    rows_to_process = min(2, len(df))
                    estimated_time = calculate_processing_time(rows_to_process)
                    st.info(f"⏱️ Estimated processing time for first {rows_to_process} rows (testing mode): {estimated_time}")
                
                # Process the data
                if process_button:
                    if not product_name or product_name == "Select a product":
                        st.error("⚠️ Please select a product before processing")
                    else:
                        # Set processing flag IMMEDIATELY to prevent double-clicks
                        st.session_state.bulk_processing = True
                        st.session_state.bulk_stop_requested = False
                        process_bulk_rca(df, selected_column, product_name)
                
                # Show results if processing is complete or in progress
                if st.session_state.get('bulk_processed_rows'):
                    show_processing_results()
                
        except Exception as e:
            st.error(f"❌ Error reading Excel file: {str(e)}")
            with st.expander("🐛 Error Details"):
                st.exception(e)
    else:
        st.info("📤 Upload an Excel file to begin")


def calculate_processing_time(num_rows: int) -> str:
    """Calculate estimated processing time based on rate limiting"""
    # Each row: 1 run_agent call (top-3 recs + BugAnalyze combined)
    # run_agent makes multiple internal tool calls, so ~20 seconds per row
    seconds_per_row = 20
    total_seconds = num_rows * seconds_per_row
    
    if total_seconds < 60:
        return f"{total_seconds} seconds"
    elif total_seconds < 3600:
        minutes = total_seconds / 60
        return f"{minutes:.1f} minutes"
    else:
        hours = total_seconds / 3600
        return f"{hours:.1f} hours"


def process_bulk_rca(df: pd.DataFrame, rca_column: str, product_name: str):
    """Process all RCAs in the DataFrame (limited to first 10 rows for testing)"""
    # Note: bulk_processing flag is set by caller before this function
    
    # Initialize results if starting fresh
    if not st.session_state.get('bulk_processed_rows'):
        st.session_state.bulk_processed_rows = []
        st.session_state.bulk_start_row = 0
    
    # Create progress indicators
    progress_bar = st.progress(0)
    status_text = st.empty()
    results_container = st.container()
    
    # **TESTING MODE: Limit to first 2 rows**
    total_rows = min(2, len(df))
    st.info(f"🧪 **Testing Mode:** Processing limited to first {total_rows} rows (out of {len(df)} total)")
    
    # Determine starting point (for resume functionality)
    start_idx = len(st.session_state.bulk_processed_rows)
    
    # Process each row
    for idx in range(start_idx, total_rows):
        # Check for stop request
        if st.session_state.get('bulk_stop_requested', False):
            st.warning("⏸️ Processing stopped by user")
            break
        
        row = df.iloc[idx]
        status_text.text(f"🔄 Processing row {idx + 1} of {total_rows}...")
        
        try:
            # Get RCA text
            rca_text = str(row[rca_column]) if pd.notna(row[rca_column]) else ""
            
            if not rca_text or rca_text.strip() == "" or rca_text == "nan":
                # Empty cell
                result = {
                    'row_number': idx + 1,
                    'rca_text_preview': "(empty)",
                    'detected_bugs': "",
                    'top3_recommendations': "",
                    'analysis_output': "",
                    'status': 'Skipped - Empty cell'
                }
            else:
                # Step 0: Detect bug IDs in RCA text (same as Page 1)
                bug_ids = list(set(re.findall(r'CSC[a-z]{2}\d{5}', rca_text, re.IGNORECASE)))
                if bug_ids:
                    detected_bugs = ", ".join(
                        f"{b.upper()} (https://cdetsng.cisco.com/webui/#view={b})"
                        for b in sorted(bug_ids)
                    )
                else:
                    detected_bugs = "None"
                
                # Step 1: Detect tech terms and score guides (mirrors Page 1 flow)
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Detecting terms and scoring guides...")
                clues_data = extract_doc_clues_data(rca_text)
                detected_terms = list(clues_data.get('term_frequencies', {}).keys())
                term_freq = clues_data.get('term_frequencies', {})
                
                # Get selected guides from bulk UI (or empty = all)
                bulk_guides = st.session_state.get('bulk_selected_guides', [])
                
                if detected_terms:
                    matched, _ = match_terms_to_guides(detected_terms, product_name, term_freq)
                    guide_scores = matched.pop('_guide_scores', {})
                    st.session_state['_guide_scores'] = guide_scores
                    st.session_state['_matched_term_guides'] = dict(matched)
                else:
                    st.session_state['_guide_scores'] = {}
                    st.session_state['_matched_term_guides'] = {}
                
                # Step 2: Build Top-3 Recommendations text (deterministic, pre-LLM)
                guide_scores = st.session_state.get('_guide_scores', {})
                matched_term_guides = st.session_state.get('_matched_term_guides', {})
                doc_inventory = load_document_inventory(product_name)
                sorted_gs = sorted(guide_scores.items(), key=lambda x: -x[1]) if guide_scores else []
                rec_lines = []
                for rank, (g_name, g_score) in enumerate(sorted_gs[:3], start=1):
                    tw = []
                    for term, guides in matched_term_guides.items():
                        if term.startswith('_'):
                            continue
                        if g_name in guides:
                            tw.append((term, 1.0 / len(guides)))
                    tw.sort(key=lambda x: -x[1])
                    section_hints = ", ".join(t.title() for t, _ in tw[:5]) if tw else "(no section hints)"
                    url = doc_inventory.get(g_name, {}).get('source_url', '')
                    url_text = url if url else '(no URL)'
                    rec_lines.append(f"#{rank}: {g_name} (score: {g_score})\n   Section hints: {section_hints}\n   URL: {url_text}")
                top3_recommendations = "\n\n".join(rec_lines) if rec_lines else "No guide scores available"
                
                # Step 3: Run run_agent with BugAnalyze.md (same as Page 1)
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Running analysis...")
                try:
                    with open("BugAnalyze.md", "r") as f:
                        question = f.read()
                except FileNotFoundError:
                    question = "Analyze the Bug/RCA content"
                
                rec_result = run_agent(product_name, question, rca_text, bulk_guides or None, detected_terms or None)
                analysis_output = rec_result['output'] if isinstance(rec_result, dict) and 'output' in rec_result else str(rec_result)
                
                # Rate limiting: Wait 10 seconds after call before next row
                time.sleep(10)
                
                result = {
                    'row_number': idx + 1,
                    'rca_text_preview': rca_text[:100] + "..." if len(rca_text) > 100 else rca_text,
                    'detected_bugs': detected_bugs,
                    'top3_recommendations': top3_recommendations,
                    'analysis_output': analysis_output,
                    'status': 'Success ✅'
                }
            
        except Exception as e:
            # Check error type
            error_str = str(e).lower()
            
            # Rate limit error (429)
            if '429' in error_str or 'rate limit' in error_str or 'spike arrest' in error_str:
                error_msg = f'Rate Limit Error: Too many API calls. System will add delays automatically.'
            # Network error
            elif any(term in error_str for term in ['connection', 'network', 'timeout', 'unreachable', 'failed to establish']):
                error_msg = f'Network Error: {str(e)[:100]} - Can resume when connection is restored'
            # Other errors
            else:
                error_msg = f'Error: {str(e)[:100]}'
            
            result = {
                'row_number': idx + 1,
                'rca_text_preview': str(row[rca_column])[:100] if pd.notna(row[rca_column]) else "(empty)",
                'detected_bugs': "",
                'top3_recommendations': "",
                'analysis_output': "",
                'status': error_msg
            }
            
            # If rate limit or network error, show alert
            if 'Rate Limit' in error_msg:
                st.warning(f"⚠️ Rate limit hit at row {idx + 1}. The system now includes 4.5s delays to prevent this. Clear results and restart for best results.")
            elif 'Network Error' in error_msg:
                st.error(f"⚠️ Network connection lost at row {idx + 1}. Your progress has been saved. Restore connection and click Resume.")
                st.session_state.bulk_stop_requested = True
        
        # Store result
        st.session_state.bulk_processed_rows.append(result)
        
        # Update progress
        progress = (idx + 1) / total_rows
        progress_bar.progress(progress)
    
    # Mark as complete
    st.session_state.bulk_processing = False
    if not st.session_state.get('bulk_stop_requested', False):
        st.session_state.bulk_processing_complete = True
        status_text.text(f"✅ Processing complete! Processed {len(st.session_state.bulk_processed_rows)} rows")
    else:
        status_text.text(f"⏸️ Processing paused at row {len(st.session_state.bulk_processed_rows)}")


def show_processing_results():
    """Display the processed results in a table and provide download"""
    st.markdown("---")
    st.markdown("#### 📊 Processing Results")
    
    results = st.session_state.bulk_processed_rows
    
    # Summary stats
    total = len(results)
    successful = sum(1 for r in results if r['status'] == 'Success ✅')
    skipped = sum(1 for r in results if 'Skipped' in r['status'])
    errors = sum(1 for r in results if 'Error' in r['status'])
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Processed", total)
    col2.metric("Successful", successful)
    col3.metric("Skipped", skipped)
    col4.metric("Errors", errors)
    
    # Results preview
    with st.expander("👀 Preview Results", expanded=True):
        preview_df = pd.DataFrame(results)
        st.dataframe(preview_df, use_container_width=True, height=400)
    
    # Prepare download
    st.markdown("#### 💾 Download Results")
    
    # Create output DataFrame by merging with original
    output_df = create_output_excel(results)
    
    # Convert to Excel with proper settings
    excel_buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl', mode='w') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Processed RCAs')
            
            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Processed RCAs']
            
            # Set column widths for better readability
            for idx, col in enumerate(output_df.columns, 1):
                # Make output columns wider
                if 'Output' in col or 'Status' in col:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 50
                else:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 20
            
            # Enable text wrapping for all cells
            from openpyxl.styles import Alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error creating Excel file: {str(e)}")
        st.info("💡 Trying simplified Excel export without formatting...")
        
        # Fallback: Create simple Excel without formatting
        excel_buffer = io.BytesIO()
        output_df.to_excel(excel_buffer, index=False, sheet_name='Processed RCAs', engine='openpyxl')
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
    
    # Download button
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bulk_rca_processed_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Download Processed Excel",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="sidebar_bulk_download_button"
    )
    
    if st.session_state.get('bulk_processing_complete'):
        st.success("✅ All rows processed successfully! You can download the results above.")
    else:
        processed_count = len(st.session_state.get('bulk_processed_rows', []))
        st.info(f"ℹ️ Processing incomplete ({processed_count} rows processed). You can download partial results above and click '▶️ Resume Processing' to continue.")


def create_output_excel(results: list) -> pd.DataFrame:
    """Create the output Excel with original data + new columns"""
    original_df = st.session_state.bulk_df.copy()
    
    # Create results dataframe
    results_df = pd.DataFrame(results)
    
    # Sanitize text fields to prevent Excel corruption
    def sanitize_text(text):
        """Clean text for Excel compatibility"""
        if pd.isna(text) or text == "":
            return ""
        text = str(text)
        # Remove null characters and other problematic characters
        text = text.replace('\x00', '').replace('\r', '\n')
        # Limit length to prevent Excel cell limit issues (32,767 characters)
        if len(text) > 32000:
            text = text[:32000] + "\n\n[Content truncated due to length]"
        return text
    
    # Apply sanitization to output columns
    if 'top3_recommendations' in results_df.columns:
        results_df['top3_recommendations'] = results_df['top3_recommendations'].apply(sanitize_text)
    if 'analysis_output' in results_df.columns:
        results_df['analysis_output'] = results_df['analysis_output'].apply(sanitize_text)
    
    # Add new columns to original dataframe
    original_df['Row_Number'] = range(1, len(original_df) + 1)
    
    # Merge results (left join to keep all original rows)
    merged_df = original_df.merge(
        results_df[['row_number', 'detected_bugs', 'top3_recommendations', 'analysis_output', 'status']],
        left_on='Row_Number',
        right_on='row_number',
        how='left'
    )
    
    # Rename columns
    merged_df = merged_df.rename(columns={
        'top3_recommendations': 'Top3_Recommendations',
        'analysis_output': 'Analysis_Output',
        'status': 'Processing_Status'
    })
    
    # Drop temporary row number column
    merged_df = merged_df.drop(columns=['row_number', 'Row_Number'])
    
    # Fill NaN in new columns (for unprocessed rows)
    merged_df['Top3_Recommendations'] = merged_df['Top3_Recommendations'].fillna('Not processed')
    merged_df['Analysis_Output'] = merged_df['Analysis_Output'].fillna('Not processed')
    merged_df['Processing_Status'] = merged_df['Processing_Status'].fillna('Not processed')
    
    return merged_df


def clear_bulk_session_state():
    """Clear all bulk analysis session state"""
    keys_to_clear = [
        'bulk_df',
        'bulk_file_name',
        'bulk_processed_rows',
        'bulk_processing',
        'bulk_processing_complete',
        'bulk_stop_requested',
        'bulk_start_row'
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]


def render_bug_section(product_name: str):
    """Render the Bug section with file upload and processing"""
    from bug2 import create_auth, get_bug_summary, get_note_content, get_all_notes, get_bug_field_values
    import xml.etree.ElementTree as ET
    
    # Important note about Streamlit behavior
    if 'bulk_bug_df' not in st.session_state:
        st.info("📌 **Note:** Upload an Excel file with bug numbers for bulk analysis")
    
    # Checkbox for note extraction
    extract_all_bug_notes = st.checkbox(
        "📋 Extract all notes (default: Behavior-changed + Release-note)",
        value=False,
        help="Check to extract all notes from bugs. Default extracts only Behavior-changed and Release-note notes, plus the Documentation-link field.",
        key="sidebar_bulk_bug_extract_all_notes"
    )
    
    # File uploader with unique key
    uploaded_file = st.file_uploader(
        "Upload Excel file containing Bug Numbers",
        type=['xlsx', 'xls'],
        help="Upload an Excel file with a column containing bug numbers",
        key="sidebar_bulk_bug_excel_uploader"
    )
    
    if uploaded_file is not None:
        try:
            # Read the Excel file only if it's a new file
            if 'bulk_bug_df' not in st.session_state or st.session_state.get('bulk_bug_file_name') != uploaded_file.name:
                df = pd.read_excel(uploaded_file)
                st.session_state.bulk_bug_df = df
                st.session_state.bulk_bug_file_name = uploaded_file.name
                st.session_state.bulk_bug_processed_rows = []
                st.session_state.bulk_bug_processing_complete = False
            else:
                # Use cached dataframe
                df = st.session_state.bulk_bug_df
            
            st.success(f"✅ File loaded: {uploaded_file.name} ({len(df)} rows, {len(df.columns)} columns)")
            
            # Show preview
            with st.expander("👀 Preview first 5 rows", expanded=True):
                st.dataframe(df.head(), use_container_width=True)
            
            # Column selection
            st.markdown("##### Select Bug Number Column")
            columns = df.columns.tolist()
            
            # Smart default: Try to find column with bug numbers (CSCw format)
            default_idx = 0
            max_bug_count = 0
            
            # First, check column names
            for idx, col in enumerate(columns):
                col_lower = str(col).lower()
                if 'bug' in col_lower or 'csc' in col_lower or 'defect' in col_lower:
                    default_idx = idx
                    break
            
            # Then, check actual cell values for CSC pattern (more reliable)
            import re
            bug_pattern = re.compile(r'^CSC[a-z]{2}\d+', re.IGNORECASE)
            
            for idx, col in enumerate(columns):
                try:
                    # Count how many cells match the bug pattern
                    bug_count = df[col].astype(str).apply(lambda x: bool(bug_pattern.match(x.strip()))).sum()
                    if bug_count > max_bug_count:
                        max_bug_count = bug_count
                        default_idx = idx
                except:
                    continue
            
            selected_column = st.selectbox(
                "Which column contains the bug numbers?",
                options=columns,
                index=default_idx,
                help="Select the column that contains bug numbers (e.g., CSCwp05354)",
                key="sidebar_bulk_bug_column_selector"
            )
            
            # Validation
            if selected_column:
                non_empty = df[selected_column].notna().sum()
                st.info(f"📊 Column '{selected_column}' has {non_empty} non-empty cells out of {len(df)} rows")
                
                # Show sample from selected column
                with st.expander("📄 Sample bug numbers from selected column"):
                    sample_bugs = df[selected_column].dropna().head(3)
                    for idx, bug in enumerate(sample_bugs, 1):
                        st.markdown(f"**Sample {idx}:** {bug}")
                
                # Processing controls
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 1, 1])
                
                with col1:
                    # Determine button text based on state
                    has_partial_results = len(st.session_state.get('bulk_bug_processed_rows', [])) > 0
                    is_complete = st.session_state.get('bulk_bug_processing_complete', False)
                    
                    if has_partial_results and not is_complete:
                        button_label = "▶️ Resume Processing"
                    else:
                        button_label = "🚀 Start Processing"
                    
                    process_button = st.button(
                        button_label,
                        type="primary",
                        use_container_width=True,
                        disabled=st.session_state.get('bulk_bug_processing', False),
                        key="sidebar_bulk_bug_process_button"
                    )
                
                with col2:
                    if st.session_state.get('bulk_bug_processing', False):
                        stop_button = st.button(
                            "⏸️ Stop",
                            type="secondary",
                            use_container_width=True,
                            key="sidebar_bulk_bug_stop_button"
                        )
                        if stop_button:
                            st.session_state.bulk_bug_stop_requested = True
                
                with col3:
                    clear_button = st.button(
                        "🗑️ Clear",
                        use_container_width=True,
                        key="sidebar_bulk_bug_clear_button"
                    )
                    if clear_button:
                        clear_bulk_bug_session_state()
                        st.rerun()
                
                # Time estimate (for first 40 rows in testing mode)
                if not st.session_state.get('bulk_bug_processing', False):
                    rows_to_process = min(2, len(df))
                    # Each bug: Fetch + 1 run_agent call (top-3 recs + BugAnalyze combined)
                    # Estimated ~20 seconds per row with delays
                    estimated_time = calculate_bug_processing_time(rows_to_process)
                    st.info(f"⏱️ Estimated processing time for first {rows_to_process} rows (testing mode): {estimated_time}")
                
                # Process the data
                if process_button:
                    if not product_name or product_name == "Select a product":
                        st.error("⚠️ Please select a product before processing")
                    else:
                        # Set processing flag IMMEDIATELY to prevent double-clicks
                        st.session_state.bulk_bug_processing = True
                        st.session_state.bulk_bug_stop_requested = False
                        process_bulk_bugs(df, selected_column, product_name, extract_all_bug_notes)
                
                # Show results if processing is complete or in progress
                if st.session_state.get('bulk_bug_processed_rows'):
                    show_bug_processing_results()
                
        except Exception as e:
            st.error(f"❌ Error reading Excel file: {str(e)}")
            with st.expander("🐛 Error Details"):
                st.exception(e)
    else:
        st.info("📤 Upload an Excel file to begin")


def calculate_bug_processing_time(num_rows: int) -> str:
    """Calculate estimated processing time for bug analysis"""
    # Each row: 1 run_agent call (top-3 recs + BugAnalyze combined)
    # run_agent makes multiple internal tool calls, so ~20 seconds per row
    seconds_per_row = 20
    total_seconds = num_rows * seconds_per_row
    
    if total_seconds < 60:
        return f"{total_seconds} seconds"
    elif total_seconds < 3600:
        minutes = total_seconds / 60
        return f"{minutes:.1f} minutes"
    else:
        hours = total_seconds / 3600
        return f"{hours:.1f} hours"


def process_bulk_bugs(df: pd.DataFrame, bug_column: str, product_name: str, extract_all_notes: bool):
    """Process all bugs in the DataFrame (limited to first 10 rows for testing)"""
    from bug2 import create_auth, get_bug_summary, get_note_content, get_all_notes, get_bug_field_values, safe_parse_cdets_xml
    from app_functions import run_agent_with_prompt_file, run_agent, match_terms_to_guides, extract_doc_clues_data, load_document_inventory
    import xml.etree.ElementTree as ET
    
    # Initialize results if starting fresh
    if not st.session_state.get('bulk_bug_processed_rows'):
        st.session_state.bulk_bug_processed_rows = []
        st.session_state.bulk_bug_start_row = 0
    
    # Create progress indicators
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # **TESTING MODE: Limit to first 2 rows**
    total_rows = min(2, len(df))
    st.info(f"🧪 **Testing Mode:** Processing limited to first {total_rows} rows (out of {len(df)} total)")
    
    # Determine starting point (for resume functionality)
    start_idx = len(st.session_state.bulk_bug_processed_rows)
    
    # Process each row
    for idx in range(start_idx, total_rows):
        # Check for stop request
        if st.session_state.get('bulk_bug_stop_requested', False):
            st.warning("⏸️ Processing stopped by user")
            break
        
        row = df.iloc[idx]
        status_text.text(f"🔄 Processing row {idx + 1} of {total_rows}...")
        
        try:
            # Get bug number
            bug_number = str(row[bug_column]).strip() if pd.notna(row[bug_column]) else ""
            
            if not bug_number or bug_number == "" or bug_number == "nan":
                # Empty cell
                result = {
                    'row_number': idx + 1,
                    'bug_number': "(empty)",
                    'top3_recommendations': "",
                    'analysis_output': "",
                    'status': 'Skipped - Empty cell'
                }
            else:
                # Step 1: Fetch bug from CDETS
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Fetching bug {bug_number} from CDETS...")
                
                auth = create_auth()
                ns = {'cdets': 'cdetsng', 'ns2': 'http://www.w3.org/1999/xlink'}
                
                # Get bug summary
                summary_response = get_bug_summary(bug_number, auth)
                summary_root = safe_parse_cdets_xml(summary_response.content)
                
                # Build bug content
                bug_content = f"# Bug {bug_number} - Complete Report\n\n"
                bug_content += "## Bug Summary\n\n"
                
                # Extract bug fields
                defect = summary_root.find('.//cdets:Defect', ns)
                if defect:
                    for field in defect.findall('.//cdets:Field', ns):
                        field_name = field.get('name')
                        field_value = field.text if field.text else 'N/A'
                        
                        if field_name in ['Headline', 'Status', 'Severity', 'Priority', 'Product', 
                                         'Component', 'Version', 'Description', 'FoundIn', 'FixedIn']:
                            bug_content += f"**{field_name}:** {field_value}\n\n"
                
                # Extract Documentation-link field
                try:
                    doc_link_values = get_bug_field_values(bug_number, 'Documentation-link', auth)
                    doc_link = doc_link_values.get('Documentation-link', 'N/A')
                    if doc_link and doc_link != 'N/A':
                        bug_content += f"**Documentation-link:** {doc_link}\n\n"
                except Exception as e:
                    # If Documentation-link field doesn't exist or error, skip silently
                    pass
                
                # Get notes
                bug_content += "\n## Notes\n\n"
                
                if extract_all_notes:
                    # Extract all notes
                    try:
                        all_note_titles = get_all_notes(bug_number, auth)
                        
                        for i, note_title in enumerate(all_note_titles, 1):
                            try:
                                note_response = get_note_content(bug_number, note_title, auth)
                                bug_content += f"### {i}. {note_title}\n\n"
                                bug_content += f"**Content:**\n{note_response.text}\n\n"
                            except Exception as e:
                                bug_content += f"*Error fetching note '{note_title}': {str(e)}*\n\n"
                    except Exception as e:
                        bug_content += f"*Error fetching notes list: {str(e)}*\n\n"
                else:
                    # Extract Behavior-changed and Release-note by default
                    default_notes = ["Behavior-changed", "Release-note"]
                    
                    for note_title in default_notes:
                        try:
                            note_response = get_note_content(bug_number, note_title, auth)
                            bug_content += f"### {note_title}\n\n"
                            bug_content += f"**Content:**\n{note_response.text}\n\n"
                        except Exception as e:
                            bug_content += f"*Note '{note_title}' not found*\n\n"
                
                # Step 2: Run top-3 recommendation workflow (same as Page 1 Analysis)
                status_text.text(f"🔄 Row {idx + 1}/{total_rows} - Running top-3 recommendation analysis...")
                
                # Detect tech terms and score guides (mirrors Page 1 flow)
                clues_data = extract_doc_clues_data(bug_content)
                detected_terms = list(clues_data.get('term_frequencies', {}).keys())
                term_freq = clues_data.get('term_frequencies', {})
                
                # Get selected guides from bulk UI (or empty = all)
                bulk_guides = st.session_state.get('bulk_selected_guides', [])
                
                if detected_terms:
                    matched, _ = match_terms_to_guides(detected_terms, product_name, term_freq)
                    guide_scores = matched.pop('_guide_scores', {})
                    st.session_state['_guide_scores'] = guide_scores
                    st.session_state['_matched_term_guides'] = dict(matched)
                else:
                    st.session_state['_guide_scores'] = {}
                    st.session_state['_matched_term_guides'] = {}
                
                # Build Top-3 Recommendations text (deterministic, pre-LLM)
                guide_scores = st.session_state.get('_guide_scores', {})
                matched_term_guides = st.session_state.get('_matched_term_guides', {})
                doc_inventory = load_document_inventory(product_name)
                sorted_gs = sorted(guide_scores.items(), key=lambda x: -x[1]) if guide_scores else []
                rec_lines = []
                for rank, (g_name, g_score) in enumerate(sorted_gs[:3], start=1):
                    # Get section hints from matched terms
                    tw = []
                    for term, guides in matched_term_guides.items():
                        if term.startswith('_'):
                            continue
                        if g_name in guides:
                            tw.append((term, 1.0 / len(guides)))
                    tw.sort(key=lambda x: -x[1])
                    section_hints = ", ".join(t.title() for t, _ in tw[:5]) if tw else "(no section hints)"
                    url = doc_inventory.get(g_name, {}).get('source_url', '')
                    url_text = url if url else '(no URL)'
                    rec_lines.append(f"#{rank}: {g_name} (score: {g_score})\n   Section hints: {section_hints}\n   URL: {url_text}")
                top3_recommendations = "\n\n".join(rec_lines) if rec_lines else "No guide scores available"
                
                # Load BugAnalyze.md as the question (same prompt as Page 1)
                try:
                    with open("BugAnalyze.md", "r") as f:
                        question = f.read()
                except FileNotFoundError:
                    question = "Analyze the Bug/RCA content"
                
                # Call run_agent with full top-3 recommendation workflow
                rec_result = run_agent(product_name, question, bug_content, bulk_guides or None, detected_terms or None)
                analysis_output = rec_result['output'] if isinstance(rec_result, dict) and 'output' in rec_result else str(rec_result)
                
                # Rate limiting: Wait 10 seconds after last call before next row
                time.sleep(10)
                
                result = {
                    'row_number': idx + 1,
                    'bug_number': bug_number,
                    'top3_recommendations': top3_recommendations,
                    'analysis_output': analysis_output,
                    'status': 'Success ✅'
                }
            
        except Exception as e:
            # Check error type
            error_str = str(e).lower()
            
            # Rate limit error (429)
            if '429' in error_str or 'rate limit' in error_str or 'spike arrest' in error_str:
                error_msg = f'Rate Limit Error: Too many API calls. System will add delays automatically.'
            # Network error
            elif any(term in error_str for term in ['connection', 'network', 'timeout', 'unreachable', 'failed to establish']):
                error_msg = f'Network Error: {str(e)[:100]} - Can resume when connection is restored'
            # Other errors
            else:
                error_msg = f'Error: {str(e)[:100]}'
            
            result = {
                'row_number': idx + 1,
                'bug_number': str(row[bug_column])[:50] if pd.notna(row[bug_column]) else "(empty)",
                'top3_recommendations': "",
                'analysis_output': "",
                'status': error_msg
            }
            
            # If rate limit or network error, show alert
            if 'Rate Limit' in error_msg:
                st.warning(f"⚠️ Rate limit hit at row {idx + 1}. The system now includes 10s delays to prevent this.")
            elif 'Network Error' in error_msg:
                st.error(f"⚠️ Network connection lost at row {idx + 1}. Your progress has been saved. Restore connection and click Resume.")
                st.session_state.bulk_bug_stop_requested = True
        
        # Store result
        st.session_state.bulk_bug_processed_rows.append(result)
        
        # Update progress
        progress = (idx + 1) / total_rows
        progress_bar.progress(progress)
    
    # Mark as complete
    st.session_state.bulk_bug_processing = False
    if not st.session_state.get('bulk_bug_stop_requested', False):
        st.session_state.bulk_bug_processing_complete = True
        status_text.text(f"✅ Processing complete! Processed {len(st.session_state.bulk_bug_processed_rows)} bugs")
    else:
        status_text.text(f"⏸️ Processing paused at row {len(st.session_state.bulk_bug_processed_rows)}")


def show_bug_processing_results():
    """Display the processed bug results in a table and provide download"""
    st.markdown("---")
    st.markdown("#### 📊 Bug Processing Results")
    
    results = st.session_state.bulk_bug_processed_rows
    
    # Summary stats
    total = len(results)
    successful = sum(1 for r in results if r['status'] == 'Success ✅')
    skipped = sum(1 for r in results if 'Skipped' in r['status'])
    errors = sum(1 for r in results if 'Error' in r['status'])
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Processed", total)
    col2.metric("Successful", successful)
    col3.metric("Skipped", skipped)
    col4.metric("Errors", errors)
    
    # Results preview
    with st.expander("👀 Preview Results", expanded=True):
        preview_df = pd.DataFrame(results)
        st.dataframe(preview_df, use_container_width=True, height=400)
    
    # Prepare download
    st.markdown("#### 💾 Download Results")
    
    # Create output DataFrame by merging with original
    output_df = create_bug_output_excel(results)
    
    # Convert to Excel with proper settings
    excel_buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl', mode='w') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Processed Bugs')
            
            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Processed Bugs']
            
            # Set column widths for better readability
            for idx, col in enumerate(output_df.columns, 1):
                # Make output columns wider
                if 'Output' in col or 'Status' in col:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 50
                else:
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = 20
            
            # Enable text wrapping for all cells
            from openpyxl.styles import Alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error creating Excel file: {str(e)}")
        st.info("💡 Trying simplified Excel export without formatting...")
        
        # Fallback: Create simple Excel without formatting
        excel_buffer = io.BytesIO()
        output_df.to_excel(excel_buffer, index=False, sheet_name='Processed Bugs', engine='openpyxl')
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
    
    # Download button
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bulk_bugs_processed_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Download Processed Excel",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        key="sidebar_bulk_bug_download_button"
    )
    
    if st.session_state.get('bulk_bug_processing_complete'):
        st.success("✅ All bugs processed successfully! You can download the results above.")
    else:
        processed_count = len(st.session_state.get('bulk_bug_processed_rows', []))
        st.info(f"ℹ️ Processing incomplete ({processed_count} bugs processed). You can download partial results above and click '▶️ Resume Processing' to continue.")


def create_bug_output_excel(results: list) -> pd.DataFrame:
    """Create the output Excel with original data + new columns"""
    original_df = st.session_state.bulk_bug_df.copy()
    
    # Create results dataframe
    results_df = pd.DataFrame(results)
    
    # Sanitize text fields to prevent Excel corruption
    def sanitize_text(text):
        """Clean text for Excel compatibility"""
        if pd.isna(text) or text == "":
            return ""
        text = str(text)
        # Remove null characters and other problematic characters
        text = text.replace('\x00', '').replace('\r', '\n')
        # Limit length to prevent Excel cell limit issues (32,767 characters)
        if len(text) > 32000:
            text = text[:32000] + "\n\n[Content truncated due to length]"
        return text
    
    # Apply sanitization to output columns
    if 'top3_recommendations' in results_df.columns:
        results_df['top3_recommendations'] = results_df['top3_recommendations'].apply(sanitize_text)
    if 'analysis_output' in results_df.columns:
        results_df['analysis_output'] = results_df['analysis_output'].apply(sanitize_text)
    
    # Add new columns to original dataframe
    original_df['Row_Number'] = range(1, len(original_df) + 1)
    
    # Merge results (left join to keep all original rows)
    merged_df = original_df.merge(
        results_df[['row_number', 'top3_recommendations', 'analysis_output', 'status']],
        left_on='Row_Number',
        right_on='row_number',
        how='left'
    )
    
    # Rename columns
    merged_df = merged_df.rename(columns={
        'top3_recommendations': 'Top3_Recommendations',
        'analysis_output': 'Analysis_Output',
        'status': 'Processing_Status'
    })
    
    # Drop temporary row number column
    merged_df = merged_df.drop(columns=['row_number', 'Row_Number'])
    
    # Fill NaN in new columns (for unprocessed rows)
    merged_df['Top3_Recommendations'] = merged_df['Top3_Recommendations'].fillna('Not processed')
    merged_df['Analysis_Output'] = merged_df['Analysis_Output'].fillna('Not processed')
    merged_df['Processing_Status'] = merged_df['Processing_Status'].fillna('Not processed')
    
    return merged_df


def clear_bulk_bug_session_state():
    """Clear all bulk bug analysis session state"""
    keys_to_clear = [
        'bulk_bug_df',
        'bulk_bug_file_name',
        'bulk_bug_processed_rows',
        'bulk_bug_processing',
        'bulk_bug_processing_complete',
        'bulk_bug_stop_requested',
        'bulk_bug_start_row'
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
